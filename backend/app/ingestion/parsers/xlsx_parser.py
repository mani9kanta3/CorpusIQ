"""
XLSX Parser using openpyxl
==========================

Why openpyxl?
-------------
1. NATIVE XLSX: Designed specifically for .xlsx files (Excel 2007+)
2. READ/WRITE: Can both read and create Excel files
3. FEATURES: Handles formulas, styles, multiple sheets
4. ACTIVE: Well-maintained, regular updates

What is an XLSX file?
---------------------
Like DOCX, XLSX is also a ZIP archive with XML inside:

    spreadsheet.xlsx (rename to .zip)
    ├── [Content_Types].xml
    ├── xl/
    │   ├── workbook.xml      # Sheet names and structure
    │   ├── worksheets/
    │   │   ├── sheet1.xml    # Actual cell data
    │   │   └── sheet2.xml
    │   ├── sharedStrings.xml # Text content (deduplicated)
    │   └── styles.xml        # Formatting
    └── docProps/
        └── core.xml          # Metadata

Installation: pip install openpyxl
"""

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pathlib import Path
from datetime import datetime
from typing import Optional

from .base import BaseParser, DocumentMetadata, ParsedDocument


class XLSXParser(BaseParser):
    """
    Parser for Microsoft Excel documents (.xlsx).
    
    Structure of parsed output:
    ---------------------------
    - content: All sheets combined as text
    - pages: Each sheet becomes one "page"
    - metadata: File info + sheet count
    
    Cell handling:
    --------------
    - Text cells: Extracted as-is
    - Number cells: Converted to string
    - Formula cells: We extract the CALCULATED VALUE, not the formula
      (e.g., if cell has "=SUM(A1:A10)" showing "150", we extract "150")
    - Empty cells: Skipped
    - Date cells: Converted to readable string
    
    Usage:
        parser = XLSXParser()
        result = parser.parse(Path("data.xlsx"))
        print(result.pages[0])  # First sheet content
    """
    
    SUPPORTED_EXTENSIONS = [".xlsx"]
    
    def parse(self, file_path: Path) -> ParsedDocument:
        """
        Parse an Excel file and extract content from all sheets.
        
        Args:
            file_path: Path to the .xlsx file
            
        Returns:
            ParsedDocument where each "page" is a worksheet
        """
        self.validate_file(file_path)
        
        try:
            # load_workbook opens the Excel file
            # data_only=True means: get calculated values, not formulas
            # Example: Cell with "=10+5" returns 15, not "=10+5"
            workbook = load_workbook(file_path, data_only=True)
            
            sheets_content = []
            
            # workbook.sheetnames gives list of sheet names: ["Sheet1", "Sales", "Summary"]
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_text = self._extract_sheet_content(sheet, sheet_name)
                sheets_content.append(sheet_text)
            
            # Combine all sheets
            full_content = "\n\n".join(sheets_content)
            
            # Extract metadata
            metadata = self._extract_metadata_from_workbook(workbook, file_path)
            
            # Close the workbook to free memory
            workbook.close()
            
            return ParsedDocument(
                content=full_content,
                metadata=metadata,
                pages=sheets_content  # Each sheet is a "page"
            )
            
        except InvalidFileException:
            raise RuntimeError(
                f"'{file_path}' is not a valid Excel file. "
                "It may be corrupted or in an older format (.xls)."
            )
        except Exception as e:
            raise RuntimeError(f"Failed to parse Excel file '{file_path}': {str(e)}")
    
    def extract_metadata(self, file_path: Path) -> DocumentMetadata:
        """
        Extract metadata from Excel file without reading all cell data.
        
        Args:
            file_path: Path to the .xlsx file
            
        Returns:
            DocumentMetadata object
        """
        self.validate_file(file_path)
        
        try:
            # read_only=True is faster for just getting metadata
            workbook = load_workbook(file_path, read_only=True)
            metadata = self._extract_metadata_from_workbook(workbook, file_path)
            workbook.close()
            return metadata
        except InvalidFileException:
            raise RuntimeError(f"'{file_path}' is not a valid Excel file.")
        except Exception as e:
            raise RuntimeError(f"Failed to extract metadata from '{file_path}': {str(e)}")
    
    def _extract_sheet_content(self, sheet, sheet_name: str) -> str:
        """
        Extract all content from a single worksheet.
        
        Format:
        -------
        Each row becomes a line, cells separated by " | "
        
        Example output:
            === Sheet: Sales Data ===
            Name | Quantity | Price | Total
            Widget A | 100 | 5.99 | 599.00
            Widget B | 50 | 12.99 | 649.50
        
        Why this format?
        ----------------
        1. Readable by humans
        2. Preserves table structure
        3. Easy for LLM to understand
        4. Can be parsed back if needed
        
        Args:
            sheet: openpyxl Worksheet object
            sheet_name: Name of the sheet (for header)
            
        Returns:
            String containing all sheet data
        """
        lines = []
        lines.append(f"=== Sheet: {sheet_name} ===")
        
        # iter_rows() iterates through all rows with data
        for row in sheet.iter_rows():
            row_values = []
            
            for cell in row:
                # cell.value can be: str, int, float, datetime, None
                value = cell.value
                
                if value is None:
                    # Empty cell - use placeholder
                    row_values.append("")
                elif isinstance(value, datetime):
                    # Format datetime nicely
                    row_values.append(value.strftime("%Y-%m-%d %H:%M:%S"))
                else:
                    # Convert to string (handles int, float, str)
                    row_values.append(str(value))
            
            # Skip completely empty rows
            if any(v.strip() for v in row_values):
                row_text = " | ".join(row_values)
                lines.append(row_text)
        
        return "\n".join(lines)
    
    def _extract_metadata_from_workbook(
        self, 
        workbook, 
        file_path: Path
    ) -> DocumentMetadata:
        """
        Extract metadata from workbook properties.
        
        Args:
            workbook: openpyxl Workbook object
            file_path: Path to the file
            
        Returns:
            DocumentMetadata object
        """
        # workbook.properties contains document metadata
        props = workbook.properties
        
        return DocumentMetadata(
            filename=file_path.name,
            file_type=self._get_file_extension(file_path),
            file_size_bytes=self._get_file_size(file_path),
            page_count=len(workbook.sheetnames),  # Number of sheets
            created_at=props.created,
            modified_at=props.modified,
            author=props.creator,  # Note: it's 'creator' not 'author' in Excel
            title=props.title
        )